"""
Parse the heterogeneous "Year wise data" MCA Excel files into one clean CSV
suitable for bulk-loading into Postgres for a company-name existence check.

Output columns: identifier, name, kind, klass, reg_date, name_norm, core_norm
 - kind:  'indian' | 'llp' | 'foreign'
 - klass: 'Private' | 'Public' | 'LLP' | 'Foreign' (best-effort)
 - name_norm: lowercase, alphanumerics only (exact-match key)
 - core_norm: name_norm with the Indian legal suffix stripped (brand key)

The files drift year to year: sheet names, header positions and column names all
vary, so we detect the header row and map columns by fuzzy token match per sheet.
Rows are de-duplicated on name_norm (a proposed-name check only needs distinct
names), keeping the earliest occurrence.
"""
import openpyxl, glob, csv, re, os, sys

# Source dir of the year-wise MCA *.xlsx files (override as argv[1]).
SRC = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\komal\Downloads\Year wise data"
OUT = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(os.path.abspath(__file__)), "mca_companies.csv")

# Brand key: normalize to alnum, then strip trailing legal-suffix tokens (longest
# first). MUST match coreKey() in controllers/mcaController.ts. NB: the final
# core_norm loaded into the DB is recomputed by project_lean.py from the name, so
# this only needs to stay consistent for standalone use of the parser output.
SUFFIX_TOKENS = sorted([
    "limitedliabilitypartnership", "onepersoncompany", "producercompany",
    "privatelimited", "publiclimited", "companylimited", "nidhilimited",
    "privateltd", "pvtlimited", "pvtltd", "section8",
    "private", "public", "limited", "company", "nidhi",
    "llp", "opc", "llc", "ltd", "pvt",
], key=len, reverse=True)

def norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())

def core(s: str) -> str:
    x = norm(s)
    changed = True
    while changed and len(x) >= 3:
        changed = False
        for tok in SUFFIX_TOKENS:
            if len(x) - len(tok) >= 3 and x.endswith(tok):
                x = x[: -len(tok)]
                changed = True
                break
    return x

def classify_sheet(name: str):
    n = name.lower()
    if "summary" in n:
        return None
    if "foreign" in n:
        return "foreign"
    if "llp" in n or "partnership" in n:
        return "llp"
    if "indian" in n or "companies reg" in n or "company" in n or "compnies" in n:
        return "indian"
    return None

def hdr_key(v) -> str:
    # Normalize underscores AND whitespace to single spaces so "COMPANY_NAME"
    # (FY2022-23) and "COMPANY NAME" (other years) map to the same key.
    return re.sub(r"[\s_]+", " ", str(v).strip().upper()) if v is not None else ""

def process():
    seen = set()
    written = 0
    per_kind = {"indian": 0, "llp": 0, "foreign": 0}
    skipped_sheets = []
    with open(OUT, "w", newline="", encoding="utf-8") as fout:
        w = csv.writer(fout)
        w.writerow(["identifier", "name", "kind", "klass", "reg_date", "name_norm", "core_norm",
                    "company_type", "activity_code", "activity_desc", "state", "address", "email"])
        for f in sorted(glob.glob(os.path.join(SRC, "*.xlsx"))):
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            for sh in wb.sheetnames:
                kind = classify_sheet(sh)
                if kind is None:
                    skipped_sheets.append((os.path.basename(f), sh, "unclassified"))
                    continue
                ws = wb[sh]
                # Materialize header scan by buffering first 25 rows.
                rows_iter = ws.iter_rows(values_only=True)
                buf = []
                for _ in range(25):
                    try:
                        buf.append(next(rows_iter))
                    except StopIteration:
                        break
                hdr_idx, colmap = find_header_in_buf(buf, kind)
                if colmap is None:
                    skipped_sheets.append((os.path.basename(f), sh, "no header found"))
                    continue
                ni = colmap["name"]
                idi = colmap.get("id")
                ci = colmap.get("class")
                di = colmap.get("date")
                ti = colmap.get("type")
                aci = colmap.get("act_code")
                adi = colmap.get("act_desc")
                si = colmap.get("state")
                addri = colmap.get("address")
                emi = colmap.get("email")

                def cell(row, idx):
                    if idx is None or idx >= len(row) or row[idx] is None:
                        return ""
                    return str(row[idx]).strip()
                # Rows after header: remaining in buf + rest of iterator.
                data_rows = buf[hdr_idx:] + list(rows_iter)
                for row in data_rows:
                    if row is None or ni >= len(row):
                        continue
                    name = row[ni]
                    if not name or not str(name).strip():
                        continue
                    name = str(name).strip()
                    if len(name) < 2:
                        continue
                    nn = norm(name)
                    if not nn or nn in seen:
                        continue
                    seen.add(nn)
                    ident = str(row[idi]).strip() if idi is not None and idi < len(row) and row[idi] else ""
                    if kind == "llp":
                        klass = "LLP"
                    elif kind == "foreign":
                        klass = "Foreign"
                    else:
                        klass = str(row[ci]).strip() if ci is not None and ci < len(row) and row[ci] else ""
                    rdate = str(row[di]).strip() if di is not None and di < len(row) and row[di] else ""
                    w.writerow([ident, name, kind, klass, rdate, nn, core(name),
                                cell(row, ti), cell(row, aci), cell(row, adi),
                                cell(row, si), cell(row, addri), cell(row, emi)])
                    written += 1
                    per_kind[kind] += 1
            wb.close()
            print(f"  done {os.path.basename(f)}  (running total {written})", flush=True)
    print("\nTOTAL written:", written)
    print("per kind:", per_kind)
    if skipped_sheets:
        print("\nSKIPPED / unmapped sheets:")
        for s in skipped_sheets:
            print("  ", s)
    print("\nOutput:", OUT, "size(MB):", round(os.path.getsize(OUT)/1e6, 1))

def find_header_in_buf(buf, kind):
    for r in range(len(buf)):
        row = [hdr_key(c) for c in buf[r]]
        has_name = any(("COMPANY NAME" in h) or ("LLP NAME" in h) or ("PARTNERSHIP NAME" in h) for h in row)
        joined = " | ".join(row)
        has_id = any(h in ("CIN", "LLPIN", "FCIN") for h in row)
        if has_name and (has_id or "DATE OF" in joined):
            colmap = {}
            for i, h in enumerate(row):
                if ("COMPANY NAME" in h or "LLP NAME" in h or "PARTNERSHIP NAME" in h) and "name" not in colmap:
                    colmap["name"] = i
                elif h in ("CIN", "LLPIN", "FCIN") and "id" not in colmap:
                    colmap["id"] = i
                elif h == "CLASS" and "class" not in colmap:
                    colmap["class"] = i
                elif ("DATE OF REGISTRATION" in h or "DATE OF INCORPORATION" in h
                      or h in ("DATE", "FOUNDED")) and "date" not in colmap:
                    colmap["date"] = i
                elif "COMPANY TYPE" in h and "type" not in colmap:
                    colmap["type"] = i
                elif ("ACTIVITY CODE" in h or h in ("ACTIVITY", "INDUSTRIAL ACTIVITY")) and "act_code" not in colmap:
                    colmap["act_code"] = i
                elif ("ACTIVITY DESCRIPTION" in h or h == "DESCRIPTION") and "act_desc" not in colmap:
                    colmap["act_desc"] = i
                elif h == "STATE" and "state" not in colmap:
                    colmap["state"] = i
                elif ("OFFICE ADDRESS" in h or h == "ADDRESS") and "address" not in colmap:
                    colmap["address"] = i
                elif h == "EMAIL" and "email" not in colmap:
                    colmap["email"] = i
            if "name" in colmap:
                return r + 1, colmap  # data starts after header row
    return None, None

if __name__ == "__main__":
    process()
