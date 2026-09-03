"""
Parse the state-wise "All Companies and LLP's" MCA workbooks into the lean CSV
that seed-mca-companies.mjs COPYs into `mca_companies`.

Source: one .xlsx per state/UT, each a single sheet with a uniform header:
  CIN | CompanyName | CompanyROCcode | CompanyCategory | CompanySubCategory |
  CompanyClass | AuthorizedCapital | PaidupCapital | CompanyRegistrationdate_date |
  Registered_Office_Address | Listingstatus | CompanyStatus | CompanyStateCode |
  CompanyIndian/Foreign Company | nic_code | CompanyIndustrialClassification

This is the FULL register (all vintages), unlike the year-wise incorporation
datasets the previous index was built from.

Output CSV columns (headerless, COPY order + two trailing columns the projector
uses to filter and then drops):
  identifier, name, kind, klass, company_type, reg_date, core_norm, status, state

Usage:  python parse_all_companies.py "<dir-of-xlsx>" out.csv
"""
import csv, glob, os, re, sys, datetime
import openpyxl

SRC = sys.argv[1]
OUT = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(os.path.abspath(__file__)), "all_companies_raw.csv")

# MUST stay in sync with coreKey() in controllers/mcaController.ts.
SUFFIX_TOKENS = sorted([
    "limitedliabilitypartnership", "onepersoncompany", "producercompany",
    "privatelimited", "publiclimited", "companylimited", "nidhilimited",
    "privateltd", "pvtlimited", "pvtltd", "section8",
    "private", "public", "limited", "company", "nidhi",
    "llp", "opc", "llc", "ltd", "pvt",
], key=len, reverse=True)

def norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())

def core_key(name: str) -> str:
    s = norm(name)
    changed = True
    while changed and len(s) >= 3:
        changed = False
        for tok in SUFFIX_TOKENS:
            if len(s) - len(tok) >= 3 and s.endswith(tok):
                s = s[: -len(tok)]
                changed = True
                break
    return s

LLPIN_RE = re.compile(r"^[A-Z]{3}-?\d{4}$")   # LLPIN, e.g. AAA-2769
FCIN_RE  = re.compile(r"^F\d+$")              # foreign company, e.g. F06967

def classify(ident: str, name: str, foreign_col: str) -> str:
    """'llp' | 'foreign' | 'indian', from the identifier shape with name/country fallbacks."""
    i = ident.strip().upper()
    if LLPIN_RE.match(i):
        return "llp"
    if FCIN_RE.match(i):
        return "foreign"
    if re.search(r"\bLLP\b", name.upper()):
        return "llp"
    fc = foreign_col.strip().lower()
    if fc and fc not in ("india", "91", "none"):
        return "foreign"
    return "indian"

def fmt_date(v) -> str:
    """DD/MM/YYYY, matching the format already stored in mca_companies."""
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    return s.split(" ")[0]

def cell(row, i):
    if i is None or i >= len(row) or row[i] is None:
        return ""
    return str(row[i]).strip()

HEADERS = {
    "cin": "id", "companyname": "name", "companycategory": "category",
    "companysubcategory": "subcategory", "companyclass": "class",
    "companyregistrationdate_date": "date", "companystatus": "status",
    "companystatecode": "state", "companyindian/foreign company": "foreign",
}

def main():
    files = sorted(glob.glob(os.path.join(SRC, "*.xlsx")))
    seen = set()               # name_norm — de-duplicate across states
    written = dupes = blank = 0
    per_kind = {"indian": 0, "llp": 0, "foreign": 0}
    per_status = {}
    with open(OUT, "w", newline="", encoding="utf-8") as fout:
        w = csv.writer(fout)
        for f in files:
            base = os.path.basename(f)
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            n_file = 0
            for sh in wb.sheetnames:
                it = wb[sh].iter_rows(values_only=True)
                try:
                    hdr = next(it)
                except StopIteration:
                    continue
                cmap = {}
                for i, h in enumerate(hdr):
                    key = HEADERS.get(str(h).strip().lower() if h is not None else "")
                    if key and key not in cmap:
                        cmap[key] = i
                if "name" not in cmap:
                    print(f"  !! {base}/{sh}: no CompanyName column, skipped", flush=True)
                    continue
                for row in it:
                    if row is None:
                        continue
                    name = cell(row, cmap["name"])
                    if len(name) < 2:
                        blank += 1
                        continue
                    nn = norm(name)
                    if not nn:
                        blank += 1
                        continue
                    if nn in seen:
                        dupes += 1
                        continue
                    seen.add(nn)
                    ident = cell(row, cmap.get("id"))
                    kind = classify(ident, name, cell(row, cmap.get("foreign")))
                    if kind == "llp":
                        klass = "LLP"
                    elif kind == "foreign":
                        klass = "Foreign"
                    else:
                        klass = cell(row, cmap.get("class"))
                    status = cell(row, cmap.get("status")) or "Unknown"
                    per_status[status] = per_status.get(status, 0) + 1
                    per_kind[kind] += 1
                    w.writerow([
                        ident, name, kind, klass,
                        cell(row, cmap.get("subcategory")),
                        fmt_date(row[cmap["date"]] if cmap.get("date") is not None and cmap["date"] < len(row) else None),
                        core_key(name), status, cell(row, cmap.get("state")),
                    ])
                    written += 1
                    n_file += 1
            wb.close()
            print(f"  {base}: +{n_file}  (total {written})", flush=True)

    print("\nTOTAL written:", written, "| dupes skipped:", dupes, "| blank:", blank)
    print("per kind:", per_kind)
    print("per status:", dict(sorted(per_status.items(), key=lambda kv: -kv[1])))
    print("Output:", OUT, "size(MB):", round(os.path.getsize(OUT) / 1e6, 1))

if __name__ == "__main__":
    main()
